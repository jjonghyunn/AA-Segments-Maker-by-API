# RESHAPE_contents_tier1_2_v2.0.py
# 2026-08-03  Jonghyun Park w/ Claude
#
# ╔══════════════════════════════════════════════════════════════════════════╗
# ║ 이 파일은 무엇인가 — 처음 보는 사람용 3줄 요약                              ║
# ╚══════════════════════════════════════════════════════════════════════════╝
# · 같은 폴더 `extract_data_v4.3_contents.py` 가 뽑은 추출 CSV 를 읽어,
#   보고서에 바로 붙일 수 있는 **union CSV 1개**로 정제한다.
# · Tier1(넓은 범위) 과 Tier2(좁은 범위) 를 **한 번에** 처리한다 — 예전엔 폴더·프로젝트가
#   갈려 2번 돌려야 했는데, 이제 추출은 1회고 여기서 site 별로 쓸 컬럼만 골라낸다.
# · 결과: output/_union_contents_tier1_2_<YYMMDD_HHMM>.csv
#
# ── [핵심 개념] valueN 이 뭔가 — 이걸 알아야 아래 설정이 이해된다 ──────────────
#   AA Workspace 의 contents 테이블은 "세로 = 기간, 가로 = 콘텐츠(CC_xx) 세그먼트" 구조다.
#   가로 컬럼이 왼쪽부터 value1, value2, … value17 로 번호가 매겨지고,
#   **어느 번호가 어느 콘텐츠인지는 5개 테이블(reportlet) 전부 동일**하다.
#     value1  = Campaign Main (기준행)      value2  = CC_00. Content Total
#     value4  = CC_02                       value5~8 = CC_03 및 하위 01/02/03
#     value10~16 = CC_04 ~ CC_10            value3/9/17 = 빈 자리(No Data)
#   → 그래서 "이 site 는 CC_03 만 쓴다" 를 **valueN 번호 목록**으로 표현할 수 있다.
#     이게 아래 TIER2_VALUE_N / SITE_VALUE_N_OVERRIDES 의 정체다.
#
# ── [site 3분류] 어느 site 가 어느 컬럼을 쓰나 ────────────────────────────────
#   ① Tier1  (TIER1_SITES 에 등록) : 전체 valueN + 콘텐츠×국가 매트릭스 필터 적용
#   ② Tier2  (그 외 전부)          : TIER2_VALUE_N (=CC_03 계열) 만
#   ③ 예외   (SITE_VALUE_N_OVERRIDES): site 별로 직접 지정 — 예) be = CC_10 만
#
# ── [값 0 처리] "행을 지우는 것"과 "값만 0으로 만드는 것"은 다르다 ─────────────
#   · valueN 필터        → **행 자체가 안 나옴** (그 site 가 안 쓰는 콘텐츠)
#   · 매트릭스 False     → 행은 나오되 **value_fx 만 0**, `value_orig` 에는 실측값 보존
#   · App 미론치(app_O_X) → 위와 동일 (value_fx 만 0, 원본 보존)
#   원본을 남기는 이유: 나중에 "이 나라 이 콘텐츠 실제로 몇이었나" 를 되짚을 수 있어야 해서다.
#
# ── 입력 형식 2종 (둘 다 자동 인식) ───────────────────────────────────────────
#   · 신규 stack_data_extract_<site>_<YYMMDD_HHMM>.csv   ← extract_data_v4.3_contents.py
#   · 구버전 column_mapping_<site>_<YYMMDD_HHMM>.csv     ← 옛 extract_data_v3.2_contents.py
#   site 별로 **가장 최신 ts 파일 1개씩**만 쓴다 (site 마다 추출 시각이 달라도 됨).
#
# 변경 이력:
#   v2.0 (2026-08-03) — Tier1/Tier2 통합 + 신규 stack 입력 대응.
#     · 입력에 stack_data_extract_* 추가 (구 column_mapping_* 도 계속 인식)
#     · TIER1_SITES / TIER2_VALUE_N / SITE_VALUE_N_OVERRIDES 로 site 별 valueN 선택
#     · 출력 컬럼명을 보고서 raw 시트와 동일하게 (tier/site_code/value_fx/value_orig …)
#     · 매트릭스 xlsx 좌표·시트명·site 보정을 전부 상단 상수로 승격
#     · 출력 1개로 통합 (_union_contents_tier1_2_*)
#   v1.2 (2026-05-29) — 콘텐츠×국가 매트릭스 필터 + __zero_fx__ (원본 보존형 0 처리)
#   v1.1 (2026-05-26) — 출력 직전 site_code 정규화 (us_old → us)
#   v1.0 (2026-05-18) — initial
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from site_registry import lookup_site

# ════════════════════════════════════════════════════════════════════
# 사용자가 바꿔야 하는 부분
# ════════════════════════════════════════════════════════════════════

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR  = SCRIPT_DIR / "output"      # 추출 CSV 가 있는 곳
OUTPUT_DIR = SCRIPT_DIR / "output"      # union 저장 위치

CURRENCY_CSV = SCRIPT_DIR / "currency.csv"    # site 별 환율 (revenue 환산용)
APP_OX_CSV   = SCRIPT_DIR / "app_O_X.csv"     # App 론치 O/X

# 출력 파일명 → _union_contents_tier1_2_<YYMMDD_HHMM>.csv
OUTPUT_BASENAME = "_union_contents_tier1_2"

# ─── ① Tier1 site — 전체 valueN + 매트릭스 필터 ─────────────────────
# 여기 없는 site 는 전부 Tier2 로 처리된다 (아래 TIER2_VALUE_N 만 사용).
# ※ be 가 Tier1 에 있는 건 의도다 — Tier1 경로로 들어가 매트릭스 필터를 받되,
#   실제로 쓰는 컬럼은 아래 SITE_VALUE_N_OVERRIDES 가 CC_10 하나로 좁힌다.
TIER1_SITES: list[str] = [
    "br", "de", "es", "in", "mx", "tr", "uk", "us", "us_old",
    "it", "pt", "au", "fr", "be", "hq",
]

# ─── ② valueN 선택 ──────────────────────────────────────────────────
# 숫자 = 위 [핵심 개념] 의 컬럼 번호. 빈 리스트 [] = 전체 통과.
# (value3/9/17 처럼 'No Data' 자리는 어차피 자동으로 걸러지므로 안 적어도 된다)
TIER1_VALUE_N: list[int] = []               # Tier1 = 전체
TIER2_VALUE_N: list[int] = [1, 5, 6, 7, 8]  # Tier2 = 기준행 + CC_03 및 하위 01/02/03

# ─── ③ site 별 valueN 예외 ──────────────────────────────────────────
# 위 ①②로 안 떨어지는 site 를 여기서 직접 지정한다. tier 판정보다 **우선** 적용.
# 예) be = "CC_10 배너만 집계" 대상 → 기준행(value1) + CC_10(value16) 만.
#     (예전 sites_input.csv 의 only_cc_10_ssd_banner 플래그를 코드로 옮긴 것)
SITE_VALUE_N_OVERRIDES: dict[str, list[int]] = {
    "be": [1, 16],
}

# ─── 출력 TIER 컬럼 라벨 ────────────────────────────────────────────
# 출력 tier 컬럼에 'Tier 1' / 'Tier 2' 를 자동으로 채운다.
#
# ⚠ 위 TIER1_SITES(=처리 경로) 와 **다른 축**이다. 헷갈리기 쉬우니 목록을 따로 둔다.
#     · TIER1_SITES       = "어느 valueN 을 쓰고 매트릭스를 적용할까" (데이터 처리)
#     · TIER1_LABEL_SITES = "보고서상 몇 티어 나라인가"                (사업 분류)
#   실제로 be 는 처리상 Tier1 경로를 타지만 사업 tier 는 Tier 2 라 두 목록이 갈린다.
#
# TIER_LABEL_ENABLED = False 면 tier 컬럼을 빈 값으로 둔다 (예전 동작 —
# 나중에 tier_subs 시트로 수기 채움).
TIER_LABEL_ENABLED = True
TIER1_LABEL_SITES: list[str] = [
    "br", "de", "es", "in", "mx", "tr", "uk", "us", "us_old",
    "it", "pt", "au", "fr", "hq",     # ← be 없음 (사업 tier 는 Tier 2)
]
TIER1_LABEL = "Tier 1"
TIER2_LABEL = "Tier 2"

# subs 컬럼은 캠페인마다 명칭 디테일이 달라져 **비워둔다** (필요하면 나중에 수기/시트 매핑).

# ─── 콘텐츠 × 국가 매트릭스 (Tier1 전용 필터) ───────────────────────
# "이 나라에서 이 콘텐츠를 실제로 노출했나" 표. False 면 그 행 value_fx 를 0 으로 만든다.
#
# ★ 원천은 **이 폴더의 contents_by_country.csv 사본**이다.
#   원본 Excel 을 절대경로로 물지 않으므로 다른 PC·repo 에 그대로 옮겨도 동작한다.
#   표가 바뀌면 Excel 에서 새로 뽑아 이 CSV 를 **덮어쓰면** 된다
#   (아래 MATRIX_REFRESH_FROM_XLSX 를 켜면 스크립트가 대신 뽑아줄 수도 있다).
MATRIX_ENABLED = True
MATRIX_CSV     = SCRIPT_DIR / "contents_by_country.csv"   # ← 원천 (사본)

# Excel 에서 자동으로 다시 뽑을지. False(기본) = 위 CSV 사본만 사용.
#   True 로 켜면 MATRIX_XLSX 를 읽어 CSV 를 덮어쓴다. 실패해도(파일 없음 / Excel 열려있음 /
#   openpyxl 미설치) 경고만 내고 기존 CSV 로 진행한다.
MATRIX_REFRESH_FROM_XLSX = False
MATRIX_XLSX = r"C:\path\to\your\contents_matrix.xlsx"
MATRIX_SHEET    = "1_Contents by Country"
MATRIX_ROWS     = (6, 15)        # segment 가 들어있는 행 범위 (엑셀 행번호, 양끝 포함)
MATRIX_COLS     = ("F", "Q")     # site 열 범위 (엑셀 열문자, 양끝 포함)
MATRIX_NAME_COL = "C"            # segment_name 이 있는 열
# Excel 에 열이 없는 site 를 코드로 보정 — {site: {segment_name: True/False}}
# 여기 지정한 site 는 명시한 segment 만 True, 나머지 segment 는 False 로 채워진다.
# (CSV 사본에 이미 그 열이 들어있으면 손댈 필요 없다 — refresh 를 켤 때만 쓰인다)
MATRIX_SITE_DEFAULTS: dict[str, dict[str, bool]] = {
    "be": {"[CAMPAIGN] CC_10. SSD Banner": True},
}

# ─── 출력 직전 site_code 통합 ───────────────────────────────────────
# RSID 가 갈린 데이터(us_old = 구 suite / us = 신 suite)를 한 site_code 로 합친다.
# rsid / start_date / end_date 는 원본을 유지하므로 출처는 계속 구분할 수 있다.
SITE_CODE_NORMALIZE: dict[str, str] = {
    "us_old": "us",
}

# ─── 라벨 / 매핑 ────────────────────────────────────────────────────
DEVICE_LABEL: dict[str, str] = {
    "pc": "PC", "mobile": "Mobile", "app": "App", "android": "Android", "ios": "iOS",
}

# metric → REPORT NO. (보고서 상 어느 표에 들어갈 값인지)
REPORT_NO_BY_METRIC = {"visits": "Engagement by Contents"}
REPORT_NO_DEFAULT   = "Order by Contents"

# 출력 컬럼 — 보고서 raw 시트와 **같은 이름·같은 순서**로 맞춰 그대로 붙여넣을 수 있게 한다.
#   value_fx   = 집계에 쓰는 값 (환율 적용 + 0 처리 반영)
#   value_orig = 0 처리 전 실측값 — **환율 미적용 원본 통화** (검증용)
#   origin_only_delayed_value = 지연전환분만 (환율 미적용, 현지통화)
OUTPUT_HEADERS = [
    "tier", "subs", "country", "site_code", "report_no", "device_type", "metric", "item",
    "value_fx", "value_orig", "origin_only_delayed_value",
    "rsid", "start_date", "end_date", "value_n",
]

# ════════════════════════════════════════════════════════════════════
# 내부 사용 (보통 건드릴 일 없음)
# ════════════════════════════════════════════════════════════════════
RE_STACK_FILE = re.compile(r"^stack_data_extract_(.+?)_(\d{6}_\d{4})\.csv$")
RE_MAP_FILE   = re.compile(r"^column_mapping_(.+?)_(\d{6}_\d{4})\.csv$")

RE_CAMPAIGN_PREFIX = re.compile(r"\[[^\]]+\]\s*")     # '[CAMPAIGN] ' 같은 캠페인 접두
RE_TRAILING_PAREN  = re.compile(r"\s*\([^)]*\)\s*$")  # 끝 괄호 '(Visit)' 등
RE_VALUE_N         = re.compile(r"^value(\d+)$", re.IGNORECASE)

# CC_ 가 없는 컬럼(기준행)의 ITEM 라벨
ITEM_PROP_FALLBACK   = "Campaign Main Visit"
ITEM_ORDER_NON_DELAY = "Campaign Main Visit > Order (Visit)"
ITEM_ORDER_DELAY     = "Campaign Main Visit > Order (Visitor)"


def _ts_now() -> str:
    return datetime.now().strftime("%y%m%d_%H%M")


def _col_to_idx(col: str) -> int:
    """엑셀 열문자 → 0-based 인덱스. 'A'→0, 'F'→5, 'AA'→26."""
    n = 0
    for ch in col.strip().upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


# ─────────────────────────────────────────────────────────────────
# 입력 파일 찾기 — site 별 최신 1개 (신규 stack / 구 column_mapping 둘 다)
# ─────────────────────────────────────────────────────────────────
def find_latest_per_site(input_dir: Path) -> list[tuple[str, str, Path, str]]:
    """[(site, ts, path, kind)] — site 별 최신 ts 1개. kind = 'stack' | 'mapping'.
    같은 site 에 두 형식이 다 있으면 ts 가 더 최신인 쪽을 쓴다."""
    best: dict[str, tuple[str, Path, str]] = {}
    for pat, kind in ((RE_STACK_FILE, "stack"), (RE_MAP_FILE, "mapping")):
        for p in input_dir.glob("*.csv"):
            m = pat.match(p.name)
            if not m:
                continue
            site, ts = m.group(1), m.group(2)
            cur = best.get(site)
            if cur is None or ts > cur[0]:
                best[site] = (ts, p, kind)
    if not best:
        raise FileNotFoundError(
            f"{input_dir} 에 추출 CSV 가 없습니다.\n"
            f"  기대 형식: stack_data_extract_<site>_<YYMMDD_HHMM>.csv "
            f"(또는 구버전 column_mapping_<site>_<YYMMDD_HHMM>.csv)\n"
            f"  → 먼저 extract_data_v4.3_contents.py 를 돌리세요."
        )
    return sorted((s, t, p, k) for s, (t, p, k) in best.items())


def load_rows(site: str, path: Path, kind: str) -> list[dict]:
    """추출 CSV 1개 → 공통 형태 dict 리스트로 변환.

    공통 키: site_code, rsid, start_date, end_date, panel, reportlet,
             device, value_n, metric, segments, data_value
    """
    with open(path, encoding="utf-8-sig", newline="") as f:
        raw = list(csv.DictReader(f))
    if not raw:
        return []

    if kind == "mapping":
        # 구버전은 이미 '1행 = 1컬럼(value_n) 총계' 형태 → 그대로 사용
        for r in raw:
            r["data_value"] = (r.get("data_value") or "").strip()
        return raw

    # ── stack (long) ─────────────────────────────────────────────
    # 1행 = (컬럼 value_n × 디멘션 item). contents 테이블은 디멘션 행이 1개뿐이라
    # 사실상 1:1 이지만, 여러 행이면 합산해 컬럼 총계를 만든다.
    bd_cols = [c for c in raw[0] if re.match(r"^bd\d+_itemId$", c or "")]
    agg: dict[tuple, dict] = {}
    for r in raw:
        # breakdown 하위 행은 제외 — 부모 총계와 이중집계된다
        if any((r.get(c) or "").strip() for c in bd_cols):
            continue
        key = (
            (r.get("site_code") or "").strip(),
            (r.get("rsid") or "").strip(),
            (r.get("start_date") or "").strip(),
            (r.get("end_date") or "").strip(),
            (r.get("panel") or "").strip(),
            (r.get("reportlet") or "").strip(),
            (r.get("device") or "").strip(),
            (r.get("value_n") or "").strip(),
            (r.get("segments") or "").strip(),
        )
        try:
            v = float((r.get("value1") or "").strip() or 0)
        except ValueError:
            v = 0.0
        if key in agg:
            agg[key]["_sum"] += v
            continue
        agg[key] = {
            "site_code": key[0], "rsid": key[1],
            "start_date": key[2], "end_date": key[3],
            "panel": key[4], "reportlet": key[5],
            "device": key[6], "value_n": key[7], "segments": key[8],
            # metric 은 원본 표기 우선 (v3.9 부터 metric_origin 이 원본, metric 은 정제본)
            "metric": (r.get("metric_origin") or r.get("metric") or "").strip(),
            "_sum": v,
        }
    out = []
    for d in agg.values():
        d["data_value"] = repr(d.pop("_sum"))
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────
# 콘텐츠 × 국가 매트릭스
# ─────────────────────────────────────────────────────────────────
def refresh_matrix_csv() -> None:
    """(옵션) 원본 Excel → CSV 사본 갱신. 실패해도 죽지 않고 기존 사본을 쓴다.

    기본은 **끄여 있다**(MATRIX_REFRESH_FROM_XLSX=False) — 폴더 안 CSV 사본이 원천이라
    외부 경로에 의존하지 않는다. 켜면 Excel 의 계산된 값만 읽어(data_only=True) CSV 를 덮어쓴다."""
    if not (MATRIX_ENABLED and MATRIX_REFRESH_FROM_XLSX):
        return
    try:
        import openpyxl
    except ImportError:
        print("[matrix] openpyxl 없음 → 기존 CSV 재사용 (pip install openpyxl 하면 자동 갱신)")
        return
    src = Path(MATRIX_XLSX)
    if not src.exists():
        print(f"[matrix] 원본 xlsx 없음 → 기존 CSV 재사용: {src}")
        return
    try:
        wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    except PermissionError:
        print("[matrix] xlsx 가 열려 있어 읽기 실패 → 기존 CSV 재사용 (Excel 을 닫으면 갱신됨)")
        return
    except Exception as e:
        print(f"[matrix] xlsx 읽기 실패({e.__class__.__name__}) → 기존 CSV 재사용")
        return
    if MATRIX_SHEET not in wb.sheetnames:
        print(f"[matrix] 시트 '{MATRIX_SHEET}' 없음 (있는 시트: {wb.sheetnames}) → 기존 CSV 재사용")
        return

    ws = wb[MATRIX_SHEET]
    r0, r1 = MATRIX_ROWS
    c0, c1 = _col_to_idx(MATRIX_COLS[0]), _col_to_idx(MATRIX_COLS[1])
    name_i = _col_to_idx(MATRIX_NAME_COL)
    grid = [list(row) for row in ws.iter_rows(min_row=1, max_row=r1,
                                              max_col=max(c1, name_i) + 1, values_only=True)]

    # site 헤더 = segment 행 바로 위 행
    hdr = grid[r0 - 2] if r0 - 2 < len(grid) else []
    sites = [str(hdr[i]).strip().lower() for i in range(c0, c1 + 1)
             if i < len(hdr) and hdr[i] not in (None, "")]
    rows_out = []
    for ri in range(r0 - 1, min(r1, len(grid))):
        row = grid[ri]
        name = row[name_i] if name_i < len(row) else None
        if not name or not str(name).strip():
            continue
        vals = [str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                for i in range(c0, c0 + len(sites))]
        rows_out.append([str(name).strip()] + vals)
    if not rows_out:
        print("[matrix] xlsx 에서 읽은 행이 0 → 기존 CSV 재사용")
        return

    # Excel 에 열이 없는 site 보정 (MATRIX_SITE_DEFAULTS)
    for site, spec in MATRIX_SITE_DEFAULTS.items():
        if site in sites:
            continue
        sites.append(site)
        want = {k.strip().lower(): bool(v) for k, v in spec.items()}
        for r in rows_out:
            r.append("True" if want.get(r[0].strip().lower(), False) else "False")

    with MATRIX_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["segment_name"] + sites)
        w.writerows(rows_out)
    print(f"[matrix] xlsx → {MATRIX_CSV.name} 갱신 ({len(rows_out)} segment × {len(sites)} site)")


def load_matrix() -> dict[tuple[str, str], bool]:
    """(정규화된 segment 이름, site) → True/False. 빈 셀은 등록 안 함(=필터 미적용)."""
    out: dict[tuple[str, str], bool] = {}
    if not MATRIX_ENABLED:
        return out
    if not MATRIX_CSV.exists():
        print(f"[matrix] {MATRIX_CSV.name} 없음 → 매트릭스 필터 비활성")
        return out
    with MATRIX_CSV.open(encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        header = next(r)
        sites = [h.strip().lower() for h in header[1:]]
        for row in r:
            if not row or not row[0].strip():
                continue
            seg = clean_segment_name(row[0])
            for site, raw in zip(sites, row[1:]):
                v = raw.strip().lower()
                if v in ("true", "1", "y", "yes"):
                    out[(seg, site)] = True
                elif v in ("false", "0", "n", "no"):
                    out[(seg, site)] = False
    return out


def clean_segment_name(name: str) -> str:
    """매트릭스 매칭용 이름 정규화 — 표 1행이 모든 변종을 잡도록 4단계로 깎는다.
       (1) 끝 괄호 반복 제거   '… (Visit)', '… (Delayed Purchase)'
       (2) US_CC_ → CC_        미국 전용 세그도 같은 CC 룰 적용
       (3) breakdown 꼬리 제거 'CC_03 … - 01. Smart Runners' → 'CC_03 …'
       (4) 소문자화
    """
    s = (name or "").strip()
    while RE_TRAILING_PAREN.search(s):
        s = RE_TRAILING_PAREN.sub("", s).strip()
    s = re.sub(r"\bUS_CC_", "CC_", s)
    s = re.sub(r"\s*-\s*\d+\.\s*.+$", "", s).strip()
    return s.lower()


# ─────────────────────────────────────────────────────────────────
# 부가 입력
# ─────────────────────────────────────────────────────────────────
def load_currency_map(path: Path) -> dict[tuple[str, str], float]:
    """(site, 연도) → 환율. 헤더의 'YYYY-MM-DD' 를 **연도만** 보고 키를 만든다.
    날짜 전체를 상수로 고정하면 cutoff 폴더를 복사할 때 매칭이 깨져 조용히 환산이 빠진다."""
    out: dict[tuple[str, str], float] = {}
    if not path.exists():
        print(f"[WARN] currency.csv 없음: {path}")
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        header = next(r)
        date_cols = [(i, m.group(1)) for i, h in enumerate(header)
                     if (m := re.match(r"^(\d{4})-\d{2}-\d{2}$", h.strip()))]
        for row in r:
            if not row or not row[0].strip():
                continue
            site = row[0].strip()
            for ci, year in date_cols:
                if ci >= len(row) or not row[ci].strip():
                    continue
                try:
                    out[(site, year)] = float(row[ci].strip())
                except ValueError:
                    pass
    return out


def load_app_x_sites(path: Path) -> set[str]:
    """App 미론치(X) site 집합. 이 site 의 app/android/ios 값은 0 으로 표시된다."""
    out: set[str] = set()
    if not path.exists():
        print(f"[WARN] app_O_X.csv 없음: {path}")
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            site = (row.get("site_code") or "").strip()
            flag = (row.get("App 론치 (O/X)") or "").strip().upper()
            if site and flag == "X":
                out.add(site)
    return out


# ─────────────────────────────────────────────────────────────────
# 값 변환 헬퍼
# ─────────────────────────────────────────────────────────────────
def value_n_num(value_n: str) -> int | None:
    m = RE_VALUE_N.match((value_n or "").strip())
    return int(m.group(1)) if m else None


def tier_of(site: str) -> str:
    """처리 경로 판정 — 어느 valueN 을 쓰고 매트릭스를 적용할지."""
    return "tier1" if site.strip().lower() in {s.lower() for s in TIER1_SITES} else "tier2"


def tier_label(site: str) -> str:
    """출력 tier 컬럼 값 (사업 분류). TIER_LABEL_ENABLED=False 면 빈 값."""
    if not TIER_LABEL_ENABLED:
        return ""
    return TIER1_LABEL if site.strip().lower() in {s.lower() for s in TIER1_LABEL_SITES} else TIER2_LABEL


def allowed_value_n(site: str) -> list[int]:
    """이 site 가 쓸 valueN 목록. [] = 전체 통과. 우선순위: site 예외 > tier 기본."""
    sc = site.strip().lower()
    if sc in {k.lower() for k in SITE_VALUE_N_OVERRIDES}:
        for k, v in SITE_VALUE_N_OVERRIDES.items():
            if k.lower() == sc:
                return v
    return TIER1_VALUE_N if tier_of(sc) == "tier1" else TIER2_VALUE_N


def extract_item(segments: str, metric: str, reportlet: str) -> str | None:
    """세그 stack 의 마지막 토큰에서 ITEM 라벨을 뽑는다. None 이면 그 행은 버린다.
    예) '[CAMPAIGN] CC_03. Scenario: … (Visit)' → '03. Scenario: …'
        'No Data'                            → None (빈 컬럼)"""
    if not segments:
        return None
    last = segments.split(";")[-1].strip()
    if re.search(r"\bno\s*data\b", last, re.IGNORECASE):
        return None
    if "CC_" in last:
        s = RE_CAMPAIGN_PREFIX.sub("", last, count=1).strip()
        s = re.sub(r"^.*?CC_", "", s)
        while RE_TRAILING_PAREN.search(s):
            s = RE_TRAILING_PAREN.sub("", s).strip()
        return s
    # CC_ 가 없는 기준행(Campaign Main)
    if "visit" in (metric or "").lower():
        return ITEM_PROP_FALLBACK
    if "delayed" in (reportlet or "").lower():
        return ITEM_ORDER_DELAY
    return ITEM_ORDER_NON_DELAY


def device_label(device: str) -> str:
    return DEVICE_LABEL.get((device or "").strip().lower(), device or "")


def report_no(metric: str) -> str:
    return REPORT_NO_BY_METRIC["visits"] if "visit" in (metric or "").lower() else REPORT_NO_DEFAULT


def type_label(metric: str, has_delayed_pair: bool) -> str:
    m = (metric or "").lower()
    if "visit" in m:
        return "Visits"
    if "order" in m:
        return "Order+Delayed Order" if has_delayed_pair else "Order"
    if "revenue" in m:
        return "Revenue+Delayed Revenue" if has_delayed_pair else "Revenue"
    return metric or ""


def normalize_reportlet_base(reportlet: str) -> str:
    """'… - Delayed Order' → '… - Order' (본 row 와 delayed row 를 짝지을 키)"""
    return re.sub(r"-\s*Delayed\s+", "- ", reportlet or "", flags=re.IGNORECASE).strip()


def normalize_segments_for_join(segments: str) -> str:
    """짝짓기용 세그 문자열 정규화 — 각 토큰의 끝 괄호를 모두 떼어 '(Visit)' / '(Delayed
    Purchase)' 차이를 없앤다."""
    parts = []
    for p in (segments or "").split(";"):
        p = p.strip()
        while True:
            new = RE_TRAILING_PAREN.sub("", p).strip()
            if new == p:
                break
            p = new
        parts.append(p)
    return "; ".join(parts).strip()


# ─────────────────────────────────────────────────────────────────
def process() -> int:
    # 0) 매트릭스 캐시 갱신
    refresh_matrix_csv()
    matrix = load_matrix()
    print(f"[matrix] {len(matrix)} (segment × site) 엔트리 로드")

    # 1) site 별 최신 추출 CSV 1개씩
    picked = find_latest_per_site(INPUT_DIR)
    n_stack = sum(1 for *_x, k in picked if k == "stack")
    print(f"[input] {len(picked)} sites (stack {n_stack} / 구 column_mapping {len(picked)-n_stack})")
    for site, ts, p, kind in picked:
        print(f"   - {site:8} {kind:8} {p.name}")

    # 2) 부가 입력
    currency = load_currency_map(CURRENCY_CSV)
    app_x_sites = load_app_x_sites(APP_OX_CSV)
    print(f"[currency] {len(currency)} (site×year) / [app X] {len(app_x_sites)} sites")

    # 3) 전체 row load
    rows: list[dict] = []
    for site, ts, p, kind in picked:
        rows.extend(load_rows(site, p, kind))
    print(f"[load] {len(rows)} rows")

    # 4) valueN 필터 — site 3분류(Tier1 / Tier2 / 예외)에 따라 쓸 컬럼만 남긴다
    kept: list[dict] = []
    drop_by_site: dict[str, int] = defaultdict(int)
    for r in rows:
        site = (r.get("site_code") or "").strip()
        allow = allowed_value_n(site)
        if not allow:
            kept.append(r)
            continue
        vn = value_n_num(r.get("value_n") or "")
        if vn is not None and vn in allow:
            kept.append(r)
        else:
            drop_by_site[site] += 1
    if drop_by_site:
        print("[valueN filter] site 별 제외 행:")
        for s in sorted(drop_by_site):
            print(f"   - {s:8} {tier_of(s):6} allow={allowed_value_n(s) or '전체'}  제외 {drop_by_site[s]}행")
    rows = kept
    print(f"[valueN filter] {len(rows)} rows 남음")

    # 5) App 미론치 site 의 app/android/ios → VALUE 만 0 (원본은 보존)
    zero_cnt = 0
    for r in rows:
        if ((r.get("site_code") or "").strip().lower() in app_x_sites
                and (r.get("device") or "").strip().lower() in {"app", "android", "ios"}):
            r["__zero_fx__"] = True
            zero_cnt += 1
    print(f"[app X zero] {zero_cnt} rows (value_fx 만 0, 원본 유지)")

    # 6) 매트릭스 False → VALUE 만 0
    #    us_old 는 us 의 매트릭스를 빌려 쓴다. 매트릭스에 없는 CC 는 건드리지 않는다.
    m_zero = m_nomatch = 0
    for r in rows:
        site = (r.get("site_code") or "").strip().lower()
        m_site = SITE_CODE_NORMALIZE.get(site, site)
        last = (r.get("segments") or "").split(";")[-1].strip()
        seg_key = clean_segment_name(last)
        if not seg_key or "cc_" not in seg_key:
            continue
        flag = matrix.get((seg_key, m_site))
        if flag is None:
            m_nomatch += 1
        elif flag is False:
            r["__zero_fx__"] = True
            m_zero += 1
    print(f"[matrix zero] {m_zero} rows (value_fx 만 0) / 매트릭스 미등록 CC row {m_nomatch}건은 통과")

    # 7) Delayed 짝 index — '… - Delayed Order/Revenue' 행을 본 행에 붙이기 위한 lookup
    delayed_index: dict[tuple, dict] = {}
    for r in rows:
        rep = r.get("reportlet") or ""
        if not re.search(r"-\s*Delayed\s+(Order|Revenue)\s*$", rep, re.IGNORECASE):
            continue
        delayed_index[(
            (r.get("site_code") or "").strip().lower(),
            (r.get("device") or "").strip().lower(),
            (r.get("panel") or "").strip(),
            normalize_reportlet_base(rep),
            (r.get("value_n") or "").strip(),
            normalize_segments_for_join(r.get("segments") or ""),
        )] = r

    # 8) 출력 row 생성
    country_cache: dict[str, str] = {}

    def country_of(sc: str) -> str:
        if sc not in country_cache:
            try:
                country_cache[sc] = lookup_site(sc).country or ""
            except Exception:
                country_cache[sc] = ""
        return country_cache[sc]

    out_rows: list[dict] = []
    delayed_used: set[tuple] = set()
    missing_rate: set[str] = set()

    for r in rows:
        site      = (r.get("site_code") or "").strip()
        device    = (r.get("device") or "").strip()
        metric    = r.get("metric") or ""
        reportlet = r.get("reportlet") or ""
        segments  = r.get("segments") or ""
        start_date = (r.get("start_date") or "").strip()
        end_date   = (r.get("end_date") or "").strip()
        rsid       = (r.get("rsid") or "").strip()
        value_n    = (r.get("value_n") or "").strip()
        zero_fx    = bool(r.get("__zero_fx__"))

        # delayed 행 자체는 본 행에 합쳐지므로 단독으로는 출력하지 않는다
        if re.search(r"-\s*Delayed\s+(Order|Revenue)\s*$", reportlet, re.IGNORECASE):
            continue

        item = extract_item(segments, metric, reportlet)
        if item is None:
            continue

        try:
            origin_val = float((r.get("data_value") or "").strip() or 0)
        except ValueError:
            origin_val = 0.0

        pair_key = (
            site.lower(), device.lower(), (r.get("panel") or "").strip(),
            normalize_reportlet_base(reportlet), value_n,
            normalize_segments_for_join(segments),
        )
        delayed_row = delayed_index.get(pair_key)
        delayed_val_only = None
        delayed_zero_fx = False
        if delayed_row is not None and re.search(r"-\s*(Order|Revenue)\s*$", reportlet, re.IGNORECASE):
            try:
                delayed_val_only = float(delayed_row.get("data_value") or 0)
            except ValueError:
                delayed_val_only = 0.0
            delayed_zero_fx = bool(delayed_row.get("__zero_fx__"))
            delayed_used.add(pair_key)

        # 환율 — 금액(revenue) 행만. 그 행의 end_date 연도로 조회한다.
        rate = 1.0
        if "revenue" in metric.lower():
            year = end_date[:4] if end_date else ""
            rate = currency.get((site.lower(), year)) or currency.get((site, year)) or 0.0
            if not rate:
                missing_rate.add(f"{site}/{year}")
                rate = 1.0

        base = {
            # tier = 사업 분류 라벨(TIER1_LABEL_SITES 기준). subs 는 캠페인마다 명칭이 달라 빈 값.
            "tier": tier_label(site), "subs": "",
            "country": country_of(site),
            "site_code": site,
            "report_no": report_no(metric),
            "device_type": device_label(device),
            "item": item,
            "rsid": rsid, "start_date": start_date, "end_date": end_date,
            "value_n": value_n,
        }

        if delayed_val_only is not None:
            out_rows.append({**base,
                             "metric": type_label(metric, has_delayed_pair=False),
                             "value_fx": 0 if zero_fx else origin_val * rate,
                             "value_orig": origin_val,
                             "origin_only_delayed_value": ""})
            item_summed = item
            last_seg = segments.split(";")[-1].strip() if segments else ""
            if "CC_" not in last_seg and "visit" not in metric.lower():
                item_summed = ITEM_ORDER_DELAY
            summed_origin = origin_val + delayed_val_only
            # 합산 행은 본 행과 delayed 행이 **둘 다** 0 대상일 때만 0
            summed_zero = zero_fx and delayed_zero_fx
            out_rows.append({**base,
                             "metric": type_label(metric, has_delayed_pair=True),
                             "item": item_summed,
                             "value_fx": 0 if summed_zero else summed_origin * rate,
                             "value_orig": summed_origin,
                             "origin_only_delayed_value": delayed_val_only})
        else:
            out_rows.append({**base,
                             "metric": type_label(metric, has_delayed_pair=False),
                             "value_fx": 0 if zero_fx else origin_val * rate,
                             "value_orig": origin_val,
                             "origin_only_delayed_value": ""})

    # 환율을 못 찾아 1.0 이 적용된 경우는 조용히 넘기지 않고 반드시 드러낸다
    if missing_rate:
        print(f"\n[WARN] 환율 미매칭 {len(missing_rate)}건 → rate=1.0 적용 (현지통화가 USD 인 척 나갑니다)")
        for k in sorted(missing_rate):
            print(f"   - {k}")
        print(f"   currency.csv 에 해당 site 행 / 연도 컬럼이 있는지 확인하세요.")

    # 9) site_code 통합 (us_old → us)
    norm_by_src: dict[str, int] = defaultdict(int)
    for r in out_rows:
        sc = r.get("site_code", "")
        new = SITE_CODE_NORMALIZE.get(sc)
        if new and new != sc:
            r["site_code"] = new
            norm_by_src[sc] += 1
    if norm_by_src:
        detail = ", ".join(f"{s}→{SITE_CODE_NORMALIZE[s]}({n})" for s, n in norm_by_src.items())
        print(f"[normalize] site_code: {detail}")

    # 10) 저장
    NUM_COLS = {"value_fx", "value_orig", "origin_only_delayed_value"}

    def _fmt(v):
        if v == "" or v is None:
            return ""
        try:
            fv = float(v)
        except (TypeError, ValueError):
            return v
        return int(fv) if fv == int(fv) else round(fv, 2)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{OUTPUT_BASENAME}_{_ts_now()}.csv"
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_HEADERS, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in out_rows:
            for c in NUM_COLS:
                r[c] = _fmt(r.get(c, ""))
            w.writerow(r)

    n_t1 = sum(1 for r in out_rows if tier_of(r["site_code"]) == "tier1")
    print(f"\n[save] {out_path}")
    print(f"  output rows : {len(out_rows)}  (tier1 {n_t1} / tier2 {len(out_rows)-n_t1})")
    print(f"  delayed pair used : {len(delayed_used)}")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    sys.exit(process())
